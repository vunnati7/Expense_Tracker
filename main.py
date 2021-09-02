import tkinter as tk
from tkinter import *
from tkinter import ttk
from PIL import ImageTk,Image
import openpyxl
import pathlib
from tkcalendar import DateEntry


m =tk.Tk()
m.title("Expense Tracker")
m.geometry("500x500")
m.resizable(0,0)
image1 = Image.open("r1.jpg")
test = ImageTk.PhotoImage(image1)
label1 = tk.Label(image=test)
label1.image = test
label1.place(x=0,y=0)
T = Text(m, height=5, width=52)
l = Label(m, text="Expense Tracker")
l.config(font=("Courier", 16))
l.pack()


def addButton():
    add=Toplevel()
    add.title("Add Expenses")
    add.geometry("500x500")
    a=Label(add,text='Add your expenses')
    a.config(font=("Courier", 16))
    a.pack()
    Date=tk.StringVar()
    Purpose=tk.StringVar()
    Amount=tk.StringVar()
    Mode=tk.StringVar()

    def submit():
        u = Date.get()
        v = Purpose.get()
        w = Amount.get()
        x = Mode.get()

        d_tf = Entry(
            add,
            width=38,
            font=('Arial', 14)
        )
        d_tf.insert(0, f'{u} {v} {w} {x}')
        d_tf.pack(pady=5)


    def export():
        u = Date.get()
        v = Purpose.get()
        w = Amount.get()
        x = Mode.get()

        file = openpyxl.load_workbook("data.xlsx")
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=u)
        sheet.cell(column=2, row=sheet.max_row, value=v)
        sheet.cell(column=3, row=sheet.max_row, value=w)
        sheet.cell(column=4, row=sheet.max_row, value=x)
        file.save("data.xlsx")
        file.close()
        add.destroy()

    lab1 = tk.Label(add, text='Date',font=("Courier", 12))
    lab1.pack(padx=100,pady=15)
    date=DateEntry(add,textvariable=Date)
    date.pack(padx=0,pady=0)

    lab2=tk.Label(add,text="Purpose",font=("Courier", 12))
    lab2.pack(padx=100, pady=15)
    purpose=tk.Entry(add,textvariable=Purpose)
    purpose.pack(padx=5,pady=5)

    lab3 = tk.Label(add, text="Cash", font=("Courier", 12))
    lab3.pack(padx=100, pady=15)
    amount = tk.Entry(add, textvariable=Amount)
    amount.pack(padx=5, pady=5)

    lab4 = tk.Label(add, text="Payment Mode", font=("Courier", 12))
    lab4.pack(padx=100, pady=15)
    mode = ttk.Combobox(add, textvariable=Mode)
    mode.pack(padx=5, pady=5)
    mode['values']=("Cash","NEFT","IMPS","UPI","Cheque")

    bb=Button(add,text="Back",bg='white',activebackground='light green',command=add.destroy)
    bb.pack(side=BOTTOM,padx=10,pady=10)

    adb=Button(add,text="Export",bg='white',activebackground='light green',command=export)
    adb.pack()

    vib=Button(add,text="View",bg='white',activebackground='light green',command=submit)
    vib.pack(side=BOTTOM,padx=10,pady=10)

    add.resizable(0,0)
    add.mainloop()


addb = Button(m,text="Add Expenses",bg='white',activebackground='light green',command=addButton)
addb.pack(side=LEFT,padx=15,pady=20)
file= pathlib.Path("data.xlsx")


if file.exists():
    pass

else:
    file=openpyxl.Workbook()
    sheet=file.active
    sheet["A1"]="Date"
    sheet["B1"]="Purpose"
    sheet["C1"]="Amount"
    sheet["D1"]="Payment Mode"
    file.save("data.xlsx")

m.mainloop()